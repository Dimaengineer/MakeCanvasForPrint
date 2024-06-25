from flask import Flask, render_template, request, redirect, session, jsonify, send_file
import shutil


from requests_oauthlib import OAuth1Session
import requests
import json
import os
import shutil
import traceback
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from io import BytesIO
import time




def CreateFolder(FolderName):
    BasePath = os.path.join(os.getcwd(), 'MakePhotoForPrintWorker')
    FolderPath = os.path.join(BasePath, FolderName)
    
    if os.path.exists(FolderPath):
        shutil.rmtree(FolderPath)
    
    os.makedirs(FolderPath)


ClientKey    = "ww8i95btnyc1k862cosb9al51hfd39vn"
ClientSecret = "ofygakpksyx23woil7fwh0avk7bbsz5n"
AccessToken  = "pgvfu2r3zob3y2uqwt2q74o36nzy9ot4"
AccessSecret = "e0hf0c9jf6ib9ar8vu1g5pasjcokbpgo"


app = Flask(__name__)
app.config["SECRET_KEY"] = "secret_key"


@app.route('/', methods=['GET', 'POST'])
def MakePhotoForPrint():
    global UsersInfo
    if request.method == "POST":
        UsersInfo[len(UsersInfo.keys())+1]={}
        session["UserId"]=max(UsersInfo.keys())

        UsersInfo[session["UserId"]]['OrdersDate']=".".join(reversed(request.form["OrdersDate"].split("-")))

        DownloadOrders=request.files.get('OrdersFile')

        CreateFolder(str(session["UserId"]))
        BasePath = os.path.join(os.getcwd(), 'MakePhotoForPrintWorker')
        FolderPath = os.path.join(BasePath, str(session["UserId"]))
        with open(f'{FolderPath}/OrdersFile.xlsx', 'wb') as OrdersFile:OrdersFile.write(DownloadOrders.read())
        UsersInfo[session["UserId"]]["FileInfo"]=f'{FolderPath}/OrdersFile.xlsx'
        UsersInfo[session["UserId"]]["FolderPath"]=FolderPath

        return redirect('/create_canvas')
    else:
        if "UserId" in session:
            BasePath = os.path.join(os.getcwd(), 'MakePhotoForPrintWorker')
            os.remove(f'{BasePath}/ResultFiles{session["UserId"]}.zip')
            
        return render_template('MakePhotoForPrint.html')



@app.route('/download_canvas', methods=['GET', 'POST'])
def DownloadCanvas():
    if request.method == 'POST':
        BasePath = os.path.join(os.getcwd(), 'MakePhotoForPrintWorker')
        return send_file(f'{BasePath}/ResultFiles{session["UserId"]}.zip', as_attachment=True, download_name="ResultFiles.zip")
    else:
        return render_template('DownloadCanvas.html')



@app.route('/load_log/', methods=['GET', 'POST'])
def LoadLog():
    if session["UserId"] in UsersInfo and os.path.exists(UsersInfo[session["UserId"]]["FolderPath"]+'/log.txt'):
        with open(UsersInfo[session["UserId"]]["FolderPath"]+'/log.txt', "r") as LogFile:
            LogText=LogFile.read()
        Log=LogText.split('\n')
        if len(Log)>1:
            Procces=int(((len(Log)-1)/len(json.loads(Log[0])))*100)
        else:
            Procces=0
        return jsonify({'Log': LogText, "Procces":str(Procces)})
    else: return jsonify({'Log': "", "Procces":"100"})


@app.route('/create_canvas', methods=['GET', 'POST'])
def CreateCanvas():
    if request.method == "POST":
        return redirect('/download_canvas')
    else:
        with open(UsersInfo[session["UserId"]]["FolderPath"]+'/log.txt', "w") as Log: Log.write("")
        return render_template('CreateCanvas.html')


@app.route('/create_canvas_function/', methods=['GET', 'POST'])
def CreateCanvasFunction():
    #Get Data
    OrdersXlsx = load_workbook(filename=UsersInfo[session["UserId"]]["FileInfo"])
    OrdersXlsxSheet=OrdersXlsx[OrdersXlsx.sheetnames[0]]
    OrderIds=[]
    for SheetLine in range(1, OrdersXlsxSheet.max_row+1):
        if str(OrdersXlsxSheet['B'+str(SheetLine)].value).replace(' ', '')==UsersInfo[session["UserId"]]['OrdersDate'] and 'дора' not in str(OrdersXlsxSheet['C'+str(SheetLine)].value).replace(' ', '').lower():
            if 'dtf' in str(OrdersXlsxSheet['C'+str(SheetLine)].value).replace(' ', '').lower() or 'сувенирка' in str(OrdersXlsxSheet['C'+str(SheetLine)].value).replace(' ', '').lower():
                OrderIds.append(OrdersXlsxSheet['A'+str(SheetLine)].value)

    with open(UsersInfo[session["UserId"]]["FolderPath"]+'/log.txt', "a") as Log: Log.write(str(OrderIds))
    
    SizesXlsx = load_workbook(filename=BytesIO(requests.get("https://docs.google.com/spreadsheets/d/1mv6epUufD3Q0UTisHDygZ7qRas8w9fHVIVh2dchitdo/export?format=xlsx").content))
    SizesXlsxSheet=SizesXlsx[SizesXlsx.sheetnames[0]]

    CreateFolder(f'{session["UserId"]}/DownloadImages')

    CreateFolder(f'{session["UserId"]}/DownloadPreviewImages')

    CreateFolder(f'{session["UserId"]}/ResultFiles')

    SheetsInfo=[]
    
    for SheetLine in range(1, SizesXlsxSheet.max_row+1):
        SheetsInfo.append([str(SizesXlsxSheet['A'+str(SheetLine)].value).replace(' ', ''), SizesXlsxSheet['B'+str(SheetLine)].value, SizesXlsxSheet['C'+str(SheetLine)].value, SizesXlsxSheet['D'+str(SheetLine)].value, SizesXlsxSheet['E'+str(SheetLine)].value, SizesXlsxSheet['F'+str(SheetLine)].value, SizesXlsxSheet['G'+str(SheetLine)].value])

    MagentoInfo = OAuth1Session(ClientKey, client_secret = ClientSecret, resource_owner_key = AccessToken, resource_owner_secret = AccessSecret)
    
    NumberDownload=0
    GetInfoText=[]

    for OrderId in OrderIds:
        try:
            GetInfoText.append('Order Number '+str(OrderId)+':')
            GetInfoText.append('')

            with open(UsersInfo[session["UserId"]]["FolderPath"]+'/log.txt', "a") as Log: Log.write("\n"+str(OrderId))

            OrderInfo=MagentoInfo.get("https://pre.mfest.com.ua/rest/V1/orders?searchCriteria[filter_groups][2][filters][0][field]=increment_id&searchCriteria[filter_groups][2][filters][0][value]="+str(OrderId)+"&searchCriteria[filter_groups][2][filters][0][condition_type]=eq")

            CreateFolder(f'{session["UserId"]}/Images')

            ImagesDir=UsersInfo[session["UserId"]]["FolderPath"]+'/Images'
            SessionRequest = requests.Session()
                    
            for Info in OrderInfo.json()['items'][0]['items'][::2]:
                for ProductCount in range(int(Info['qty_ordered'])):
                    Name=Info['name']
                    FullSKU=Info['sku'].replace(' ', '')
                    
                    ElementExsist=0

                    Result = next((SheetLine for SheetLine in SheetsInfo if SheetLine[0] == FullSKU), None)

                    if Result==None:
                        pass
                    elif Result[4]=='no':
                        ElementExsist=1
                    else:
                        ElementExsist=2

                    
                    if ElementExsist==0:
                        GetInfoText.append("Ім'я: "+Name)
                        GetInfoText.append('Повний SKU: '+FullSKU)
                        GetInfoText.append('Статус: Скасовано. Цього SKU немає в списку')
                        GetInfoText.append('')
                    elif ElementExsist==1:
                        GetInfoText.append("Ім'я: "+Name)
                        GetInfoText.append('Повний SKU: '+FullSKU)
                        GetInfoText.append('Статус: Скасування. Елемент не використовується')
                        GetInfoText.append('')
                    else:
                        for Info2 in OrderInfo.json()['items'][0]['items']:
                            if 'extension_attributes' in Info2 and str(Info['item_id']) in str(Info2['extension_attributes']['design_info']['archive']) and Info2['sku'].replace(' ', '')==FullSKU:
                                #print(Info2)
                                ImageNumber=1
                                for ImageBaseUrl, ImagePrintUrl in zip([Base['url'] for Base in Info2['extension_attributes']['design_info']['images'] if '/base/' in Base['url']], [Print['url'] for Print in Info2['extension_attributes']['design_info']['images'] if '/print/' in Print['url']]):
                                    with open(UsersInfo[session["UserId"]]["FolderPath"]+'/Images/PrintImage'+str(ImageNumber)+'.'+str(ImagePrintUrl.split('.')[-1]), 'wb') as DownloadArchiveFile, SessionRequest.get(ImagePrintUrl, stream=True) as DownloadArchiveRequest:
                                        for DownloadArchiveChunk in DownloadArchiveRequest.iter_content(chunk_size=8192):
                                            DownloadArchiveFile.write(DownloadArchiveChunk)
                                    ImageNumber+=1 
                                    with open(UsersInfo[session["UserId"]]["FolderPath"]+'/Images/BaseImage'+str(ImageNumber)+'.'+str(ImageBaseUrl.split('.')[-1]), 'wb') as DownloadArchiveFile, SessionRequest.get(ImageBaseUrl, stream=True) as DownloadArchiveRequest:
                                        for DownloadArchiveChunk in DownloadArchiveRequest.iter_content(chunk_size=8192):
                                            DownloadArchiveFile.write(DownloadArchiveChunk)
                                    ImageNumber+=1
                                break
                        time.sleep(0.5)  
                        SortedImagesList=sorted([ImageNames.name.replace('Base', '').replace('Print', '') for ImageNames in list(os.scandir(ImagesDir)) if '.png' in ImageNames.name or '.jpg' in ImageNames.name ])
                        ProductWidthsHeightsSidesTypePrints=[]
                        for ImageNumber, ImageName in enumerate(SortedImagesList):
                            if ImageNumber%2==0:
                                ElementIsIncorrect=False
                                ImageSize=[]
                                for SheetLine in SheetsInfo:
                                    if SheetLine[0]==FullSKU:
                                        ImageSize.append([SheetLine[4], SheetLine[5], SheetLine[2], SheetLine[3], SheetLine[1], SheetLine[6]])


                                Picture=Image.open(ImagesDir+'/Print'+ImageName)
                                PictureWidth, PictureHeight=Picture.size
                                NeededWidth=0
                                for PrintWidth, PrintHeight, Width, Height, Side, TypePrint in ImageSize:
                                    if int(Width)==PictureWidth and int(Height)==PictureHeight:
                                        NeededWidth=round(float(str(PrintWidth).replace(',', '.'))*59.05511811)
                                        NeededHeight=round(float(str(PrintHeight).replace(',', '.'))*59.05511811)
                                        PictureSide=Side
                                        PictureTypePrint=TypePrint
                                        break
                                if NeededWidth!=0 and Picture.getbbox()!=None:
                                    if "Чашка" in FullSKU or 'Khameleon' in FullSKU:
                                        if  Picture.getbbox()[2]-Picture.getbbox()[0]<Picture.size[0]:
                                            NeededHeight=round(float(9)*59.05511811)
                                        else:
                                            NeededHeight=round(float(10)*59.05511811)

                                    TopSpace, BottomSpace, LeftSpace, RightSpace=Picture.getbbox()[1], PictureHeight-Picture.getbbox()[3], Picture.getbbox()[0], PictureWidth-Picture.getbbox()[2]
                                    TopSpace, BottomSpace, LeftSpace, RightSpace=TopSpace/PictureHeight, BottomSpace/PictureHeight, LeftSpace/PictureWidth, RightSpace/PictureWidth
                                    Picture=Picture.crop(Picture.getbbox())
                                    PictureWidth, PictureHeight=Picture.size
                                    if PictureWidth/NeededWidth>=PictureHeight/NeededHeight:
                                        NeededWidth2=round(NeededWidth-(NeededWidth*LeftSpace+NeededWidth*RightSpace))
                                        if "Носки" in FullSKU:
                                            NeededWidth2+=118
                                        Picture=Picture.resize((NeededWidth2, round(NeededWidth2/PictureWidth*PictureHeight)), Image.LANCZOS)
                                    else:
                                        NeededHeight=round(NeededHeight-(NeededHeight*TopSpace+NeededHeight*BottomSpace))
                                        NeededWidth=round(NeededHeight/PictureHeight*PictureWidth)
                                        if "Носки" in FullSKU:
                                            NeededWidth+=118
                                        Picture=Picture.resize((NeededWidth, NeededHeight), Image.LANCZOS)
                                    if '3D' not in FullSKU:
                                        Picture.save(UsersInfo[session["UserId"]]["FolderPath"]+'/DownloadImages/'+str(OrderId)+'|'+FullSKU.replace('/', '\\')+'|'+PictureSide+'|'+PictureTypePrint+"|"+str(NumberDownload)+'.png', dpi=(150, 150))
                                    else:
                                        Picture.save(UsersInfo[session["UserId"]]["FolderPath"]+'/DownloadImages3D/'+str(OrderId)+'|'+FullSKU.replace('/', '\\')+'|'+PictureSide+'|'+PictureTypePrint+"|"+str(NumberDownload)+'.png', dpi=(150, 150))

                                    ProductWidthsHeightsSidesTypePrints.append([NeededWidth, NeededHeight, PictureSide, PictureTypePrint])
                                
                                    PreviewPicture=Image.open(ImagesDir+'/Base'+SortedImagesList[ImageNumber+1])
                                    PreviewPicture=PreviewPicture.resize((160, int(160/PreviewPicture.size[0]*PreviewPicture.size[1])), Image.LANCZOS)
                                    if '3D' not in FullSKU:
                                        PreviewPicture.save(UsersInfo[session["UserId"]]["FolderPath"]+'/DownloadPreviewImages/'+str(OrderId)+'|'+FullSKU.replace('/', '\\')+'|'+PictureSide+'|'+PictureTypePrint+"|"+str(NumberDownload)+'.png')
                                    else:
                                        PreviewPicture.save(UsersInfo[session["UserId"]]["FolderPath"]+'/DownloadPreviewImages3D/'+str(OrderId)+'|'+FullSKU.replace('/', '\\')+'|'+PictureSide+'|'+PictureTypePrint+"|"+str(NumberDownload)+'.png')

                                    if "Носки" in FullSKU:
                                        NumberDownload+=1
                                        Picture.save(UsersInfo[session["UserId"]]["FolderPath"]+'/DownloadImages/'+str(OrderId)+'|'+FullSKU.replace('/', '\\')+'|'+PictureSide+'|'+PictureTypePrint+"|"+str(NumberDownload)+'.png', dpi=(150, 150))
                                        PreviewPicture.save(UsersInfo[session["UserId"]]["FolderPath"]+'/DownloadPreviewImages/'+str(OrderId)+'|'+FullSKU.replace('/', '\\')+'|'+PictureSide+'|'+PictureTypePrint+"|"+str(NumberDownload)+'.png')

                                    NumberDownload+=1
                                elif Picture.getbbox()==None:
                                    GetInfoText.append("Ім'я: "+Name)
                                    GetInfoText.append('Повний SKU: '+FullSKU)
                                    GetInfoText.append('Сторона: '+Side)
                                    GetInfoText.append("Статус: Изображение в товаре пустое!")
                                    GetInfoText.append('')
                                else:
                                    ElementIsIncorrect=True

                        
                        if ElementIsIncorrect==True:
                            GetInfoText.append("Ім'я: "+Name)
                            GetInfoText.append('Повний SKU: '+FullSKU)
                            GetInfoText.append("Статус: В товаре указаны неверные данные")
                            GetInfoText.append('')
                        else:
                            for Width, Height, Side, TypePrint in ProductWidthsHeightsSidesTypePrints:
                                GetInfoText.append("Ім'я: "+Name)
                                GetInfoText.append('Повний SKU: '+FullSKU)
                                GetInfoText.append('Сторона: '+Side)
                                GetInfoText.append('Ширина зображення (см): '+str(round(int(Width)/59.05511811)))
                                GetInfoText.append('Висота зображення (см): '+str(round(int(Height)/59.05511811)))
                                GetInfoText.append('Тип печати изображения: '+str(PictureTypePrint))
                                GetInfoText.append('Статус: Створено успішно')
                                GetInfoText.append('')
                                
        except Exception as Error:
            with open(UsersInfo[session["UserId"]]["FolderPath"]+'/log.txt', "a") as Log: Log.write("\n"+str(traceback.format_exc()))

            GetInfoText.append("В продукте ошибка на втором этапе: "+str(traceback.format_exc()))
            GetInfoText.append('')


        GetInfoText.append('')
        GetInfoText.append('')

        
                    

                                
    with open(UsersInfo[session["UserId"]]["FolderPath"]+'/ResultFiles/LOG File.txt', 'w') as LOGFile: LOGFile.write('\n'.join(GetInfoText))



    #Third Step To Picture
    PictureDTF=Image.new('RGBA', (3729, 17717), color = (255, 255, 255,0))
    PictureDTFDraw=ImageDraw.Draw(PictureDTF)
    PictureSubli=Image.new('RGBA', (3872, 11811), color = (255, 255, 255,0))
    PictureSubliDraw=ImageDraw.Draw(PictureSubli)
    PictureDTFNumber=1
    PictureSubliNumber=1
    EmptySpacesDTF=[]
    EmptySpacesSubli=[]

    while True:
        if all(len(ImagePath.split('|')) > 2 for ImagePath in os.listdir(UsersInfo[session["UserId"]]["FolderPath"] + '/DownloadImages') if ".png" in ImagePath):
            break
        time.sleep(0.5)


    def CreatePicture(EmptySpacesList, FinishPictureWidth, FinishPicture, FinishPictureFullWidth, FinishPictureDraw, PictureNumber, MaxImageInARow, PictureType, FinishPictureFullHeight):
        OpenPicture=Image.open(UsersInfo[session["UserId"]]["FolderPath"]+'/DownloadImages/'+ImagePath).convert("RGBA")
        OpenPreviewPicture=Image.open(UsersInfo[session["UserId"]]["FolderPath"]+'/DownloadPreviewImages/'+ImagePath).convert("RGBA")
        
        if OpenPicture.size[0]>1654:
            OpenPicture = OpenPicture.rotate(-90, expand=True)

        TextInfo=str(ImagePath.split('|')[0])+'\n'+str(ImagePath.split('|')[1])
        TextSize=str(ImagePath.split('|')[-3]).replace('front', 'Перед').replace('back', 'Спина')
        
        PictureText=Image.new('RGBA', (2000, 2000), color = (255, 255, 255,0))
        PictureTextDraw=ImageDraw.Draw(PictureText)
        
        if TextSize=='Спина':
            PictureTextDraw.text((0, 0), TextInfo, fill='black', font=ImageFont.truetype('AmericanTypewriterBold.ttf', size=30), antialias=True)
            PictureTextDraw.text((0, PictureText.getbbox()[3]-PictureText.getbbox()[1] + 7), TextSize, fill='red', font=ImageFont.truetype('AmericanTypewriterBold.ttf', size=70), antialias=True)
        else:
            PictureTextDraw.text((0, 0), TextInfo+f'\n{TextSize}', fill='black', font=ImageFont.truetype('AmericanTypewriterBold.ttf', size=30), antialias=True)
        PictureText=PictureText.crop(PictureText.getbbox())

        if PictureText.size[0]>=OpenPicture.size[0]:
            PictureText=PictureText.resize((OpenPicture.size[0], int((PictureText.size[1]*OpenPicture.size[0])/PictureText.size[0])), Image.LANCZOS)
            
        TextHeight=PictureText.size[1]
        FullPicture=Image.new('RGBA', (OpenPicture.size[0], OpenPicture.size[1]+25+TextHeight), color = (255, 255, 255,0))
        FullPicture.paste(OpenPicture, (0, 0), mask=OpenPicture)
        FullPicture.paste(PictureText, (0, OpenPicture.size[1]+25), mask=PictureText)
        

        

        UseEmptySpace=None
        for EmptySpaceIndex, EmptySpace in enumerate(reversed(EmptySpacesList)):
            EmptySpaceWidth=FinishPictureWidth-EmptySpace[0][-1][0]
            EmptySpaceHeight=EmptySpace[0][-1][1]
            if FullPicture.size[0]<=EmptySpaceWidth and FullPicture.size[1]<=EmptySpaceHeight:
                UseEmptySpace=len(EmptySpacesList)-1-EmptySpaceIndex
                break

        if UseEmptySpace==None:
            FinishPictureHeight=FinishPicture.getbbox()[3] if FinishPicture.getbbox() != None else 0

            if FinishPictureHeight+FullPicture.size[1]>FinishPictureFullHeight:
                FinishPicture=FinishPicture.crop((0, FinishPicture.getbbox()[1], FinishPictureFullWidth, FinishPicture.getbbox()[3])) 
                FinishPicture.save(UsersInfo[session["UserId"]]["FolderPath"]+f'/ResultFiles/FinishImage{PictureType}{str(PictureNumber)}.tiff', dpi=(150, 150))
                PictureNumber+=1
                FinishPicture=Image.new('RGBA', (FinishPictureFullWidth, FinishPictureFullHeight), color = (255, 255, 255,0))
                FinishPictureDraw=ImageDraw.Draw(FinishPicture)
                EmptySpacesList=[]
                FinishPictureHeight=FinishPicture.getbbox()[3] if FinishPicture.getbbox() != None else 0
            
            EmptySpacesList.append([ [( int(FullPicture.size[0])+30, int(FullPicture.size[1]) )], FinishPictureHeight])
            FinishPicture.paste(FullPicture, (0, FinishPictureHeight+60), mask=FullPicture)
            FinishPicture.paste(OpenPreviewPicture, (FinishPictureWidth+1, FinishPictureHeight+60), mask=OpenPreviewPicture)
        else:
            EmptySpaceWidth=FinishPictureWidth-EmptySpacesList[UseEmptySpace][0][-1][0]
            EmptySpaceHeight=EmptySpacesList[UseEmptySpace][0][-1][1]
            EmptyHeight=EmptySpacesList[UseEmptySpace][1]

            FinishPicture.paste(FullPicture, (EmptySpacesList[UseEmptySpace][0][-1][0], EmptyHeight+60), mask=FullPicture)
            FinishPicture.paste(OpenPreviewPicture, (FinishPictureWidth+1+(int(len(EmptySpacesList[UseEmptySpace][0]))*160), EmptyHeight+60), mask=OpenPreviewPicture)

            EmptySpacesList[UseEmptySpace][0].append((EmptySpacesList[UseEmptySpace][0][-1][0]+FullPicture.size[0]+30, EmptySpaceHeight))

            if len(EmptySpacesList[UseEmptySpace][0])==MaxImageInARow:
                EmptySpacesList=EmptySpacesList[:UseEmptySpace]+EmptySpacesList[UseEmptySpace+1:]
            
            

        return [EmptySpacesList, FinishPicture, FinishPictureDraw, PictureNumber]



    for ImagePath in sorted([ImagePath for ImagePath in os.listdir(UsersInfo[session["UserId"]]["FolderPath"]+'/DownloadImages') if ".png" in ImagePath], key=lambda x: x.split('|')[1]):
        if ImagePath.split('|')[-2]=="DTF":
            CreatedPicture=CreatePicture(EmptySpacesDTF, 3248, PictureDTF, 3729, PictureDTFDraw, PictureDTFNumber, 3, "DTF", 17717)
            EmptySpacesDTF=CreatedPicture[0]
            PictureDTF=CreatedPicture[1]
            PictureDTFDraw=CreatedPicture[2]
            PictureDTFNumber=CreatedPicture[3]
        else:
            CreatedPicture=CreatePicture(EmptySpacesSubli, 3071, PictureSubli, 3872, PictureSubliDraw, PictureSubliNumber, 5, "Subli", 11811)
            EmptySpacesSubli=CreatedPicture[0]
            PictureSubli=CreatedPicture[1]
            PictureSubliDraw=CreatedPicture[2]
            PictureSubliNumber=CreatedPicture[3]

        
    if PictureDTF.getbbox()!=None:
        PictureDTF=PictureDTF.crop((0, PictureDTF.getbbox()[1], 3729, PictureDTF.getbbox()[3])) 
        PictureDTF.save(UsersInfo[session["UserId"]]["FolderPath"]+f'/ResultFiles/FinishImageDTF{str(PictureDTFNumber)}.tiff', dpi=(150, 150))

    if PictureSubli.getbbox()!=None:
        PictureSubli=PictureSubli.crop((0, PictureSubli.getbbox()[1], 3872, PictureSubli.getbbox()[3])) 
        PictureSubli.save(UsersInfo[session["UserId"]]["FolderPath"]+f'/ResultFiles/FinishImageSubli{str(PictureSubliNumber)}.tiff', dpi=(150, 150))

    BasePath = os.path.join(os.getcwd(), 'MakePhotoForPrintWorker')        
    shutil.make_archive(f'{BasePath}/ResultFiles{session["UserId"]}', 'zip', UsersInfo[session["UserId"]]["FolderPath"]+'/ResultFiles')
    if os.path.exists(UsersInfo[session["UserId"]]["FolderPath"]):
        shutil.rmtree(UsersInfo[session["UserId"]]["FolderPath"])

    return jsonify({'Log': ""})



if __name__ == '__main__':
    UsersInfo = {}
    app.run(debug=False)